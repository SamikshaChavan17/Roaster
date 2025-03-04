import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
import random
import os
from tkcalendar import Calendar
 
# File Paths
LEAVE_FILE = r"C:\Users\SC001062379\Desktop\roaster\leave_data.xlsx"
SHIFT_FILE = r"C:\Users\SC001062379\Desktop\roaster\shift_assignments_month2.xlsx"
 
# L1, L2, and L3 Users
L1_USERS = [
    "Prasanna Gadamsetty-L1", "Gopi Gade-L1", "Mayur Gujarathi-L1", "Macharla Srinivas-L1", "Sunil-L1",
    "Aachal Thakare-L1", "Aakash Sinha-L1", "Hussain Dudekula-L1", "Paras Bharat-L1", "Mansi Arora-L1",
    "Anil-L1", "Shazeb-L1", "Bhargavi-L1"
]
L2_USERS = ["Ankadi Lokesh Manivarma-L2", "Vinod Kumar-L2", "Sourish Bhowmik-L2", "Shresht Jain-L2"]
L3_USERS = ["Naveen Kumar-L3", "Vijay Bhashanpally-L3", "Vijay Kumar Sinha-L3"]
ALL_USERS = L1_USERS + L2_USERS + L3_USERS
 
SHIFT_OPTIONS = ["M1", "A2", "N3"]  # Morning, Afternoon, Night
GENERAL_SHIFT = "G"  # General shift
POSSIBLE_WEEKOFFS = ["Sun-Mon", "Mon-Tue", "Tue-Wed", "Wed-Thu", "Thu-Fri", "Fri-Sat", "Sat-Sun"]
 
months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
 
# Function to assign week-offs
def assign_weekoffs():
    weekoff_assignments = {}
    weekoff_assignments["Paras Bharat-L1"] = ("Sat", "Sun")
    weekoff_assignments["Bhargavi-L1"] = ("Sat", "Sun")
 
    for user in ALL_USERS:
        if user not in weekoff_assignments:
            weekoff_assignments[user] = random.choice(POSSIBLE_WEEKOFFS)
 
    return weekoff_assignments
 
# Function to assign shifts
def assign_shifts_for_month(selected_month):
    assigned_shifts = {}
    weekoff_assignments = assign_weekoffs()
    year = datetime.now().year
    month_index = months.index(selected_month) + 1
    days_in_month = (datetime(year, month_index + 1, 1) - timedelta(days=1)).day
 
    for user in ["Paras Bharat-L1", "Bhargavi-L1"]:
        assigned_shifts[user] = ["G" if (datetime(year, month_index, d).strftime('%a') not in ["Sat", "Sun"]) else "W" for d in range(1, days_in_month + 1)]
 
    remaining_users = [user for user in ALL_USERS if user not in ["Paras Bharat-L1", "Bhargavi-L1"]]
    random.shuffle(remaining_users)
 
    for user in remaining_users:
        shifts_for_user = []
        current_shift = random.choice(SHIFT_OPTIONS)
        weekoff_days = weekoff_assignments[user]
 
        for day in range(1, days_in_month + 1):
            week_day = datetime(year, month_index, day).strftime('%a')
            if week_day in weekoff_days:
                shifts_for_user.append("W")
            else:
                if len(shifts_for_user) >= 2 and shifts_for_user[-1] == "W" and shifts_for_user[-2] == "W":
                    available_shifts = [s for s in SHIFT_OPTIONS if s != current_shift]
                    current_shift = random.choice(available_shifts)
                shifts_for_user.append(current_shift)
 
        assigned_shifts[user] = shifts_for_user
 
    return assigned_shifts, days_in_month
 
# Function to generate roster
def generate_monthly_roster(selected_month):
    assigned_shifts, days_in_month = assign_shifts_for_month(selected_month)
    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Assignments"
 
    headers = ["User", "User Type"] + [f"{i+1} {selected_month}" for i in range(days_in_month)] + ["Morning (M1)", "Afternoon (A2)", "Night (N3)", "General (G)", "Week-off (W)", "Leave Count(L)"]
    ws.append(headers)
 
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
 
    for user in ALL_USERS:
        user_type = "L1" if user in L1_USERS else "L2" if user in L2_USERS else "L3"
        shifts = assigned_shifts.get(user, ["Not Assigned"] * days_in_month)
        morning_count, afternoon_count, night_count, general_count, weekoff_count = shifts.count("M1"), shifts.count("A2"), shifts.count("N3"), shifts.count("G"), shifts.count("W")
        ws.append([user, user_type] + shifts + [morning_count, afternoon_count, night_count, general_count, weekoff_count, 0])
 
        wb.save(SHIFT_FILE)
 
# Function to load leave data
def load_leave_data():
    leave_data = []
    wb = load_workbook(LEAVE_FILE)
    sheet = wb.active
 
    for row in range(2, sheet.max_row + 1):
        user, leave_date_str, leave_type, status = sheet.cell(row, 1).value, sheet.cell(row, 2).value, sheet.cell(row, 3).value, sheet.cell(row, 4).value
        if user and leave_date_str and status and status.lower() == "approved":
            try:
                leave_date = datetime.strptime(leave_date_str, "%d-%m-%Y").date()
                leave_data.append((user.strip(), leave_date, leave_type))
            except ValueError:
                print(f"Skipping invalid date format for user {user}: {leave_date_str}")
 
    return leave_data
 
# Function to update roster with leave
def update_roster_with_leave():
    leave_data = load_leave_data()
    wb = load_workbook(SHIFT_FILE)
    sheet = wb.active
 
    header_row = sheet[1]
    date_column_map = {int(cell.value.split()[0]): col for col, cell in enumerate(header_row[2:-6], start=3)}
 
    shift_data = {sheet.cell(row, 1).value: row for row in range(2, sheet.max_row + 1)}
 
    leave_counts = {}
 
    for user, leave_date, leave_type in leave_data:
        if user in shift_data:
            row = shift_data[user]
            day = leave_date.day
            if day in date_column_map:
                col = date_column_map[day]
                sheet.cell(row=row, column=col, value="L")
                leave_counts[user] = leave_counts.get(user, 0) + 1
 
    for user, leave_count in leave_counts.items():
        if user in shift_data:
            sheet.cell(row=shift_data[user], column=sheet.max_column, value=leave_count)
 
    wb.save(SHIFT_FILE)
root = tk.Tk()
root.title("Leave Application System")
root.geometry("550x500")
root.configure(bg="#e3f2fd")

frame = tk.Frame(root, padx=20, pady=20, bg="#ffffff", relief=tk.GROOVE, borderwidth=5)
frame.pack(pady=20)

tk.Label(frame, text="Shift Assignment & Leave Management", font=("Arial", 16, "bold"), fg="#1a237e", bg="#ffffff").pack(pady=10)

# Month selection dropdown
month_label = tk.Label(frame, text="Select Month to Generate Roster", font=("Arial", 12, "bold"), bg="#ffffff")
month_label.pack(pady=10)

months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
month_var = tk.StringVar(value=months[datetime.now().month - 1])  # Set current month as default
month_dropdown = tk.OptionMenu(frame, month_var, *months)
month_dropdown.pack(pady=10)

# Button to generate roster based on selected month
generate_roster_btn = tk.Button(frame, text="Generate Monthly Roster", command=lambda: generate_monthly_roster(month_var.get()), bg="#0288d1", fg="white", font=("Arial", 12, "bold"), padx=10, pady=5, relief=tk.RAISED, borderwidth=3)
generate_roster_btn.pack(pady=10)

root.mainloop()